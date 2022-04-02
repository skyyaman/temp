##这个版本用在服务器端。结果放在txt文件中
import requests
import re
import json
import openpyxl
import time,random
import sys,os

startshop=123027   #25/3中午运行到这里了 1000020000
visitkey='45108802361820016'
run_times=2500
modikey=visitkey
ns = [[0 for x in range(3)] for y in range(1)]
j=0
runtimes=0
successTimes=0
tellme=0
blacktimes=0
startTime = time.time()

def create_xls(filename):
    if not os.path.isfile(filename):
        wb = openpyxl.Workbook()
        sht = wb.worksheets[0]
        sht["A1"] = '店铺'
        sht["B1"] = 'token'
        sht.title = "sheet1"
        wb.save(filename)
        wb2 = openpyxl.load_workbook(filename, keep_vba=True)
        wb2.save(filename)
def write_txt(filename,txt):
    with open(filename,"a") as f:
        f.write(txt)
def alter(file,old_str,new_str):
    file_data = ""
    replaced=0

    with open(file, "r", encoding="utf-8") as f:
        for line in f:
            if old_str in line:
                replaced+=1
                if replaced==1: line = line.replace(old_str,new_str)
            file_data += line
    with open(file,"w",encoding="utf-8") as f:
        f.write(file_data)
#create_xls('data.xlsm')
logging=str(time.strftime('%H:%M:%S', time.localtime()))+':开始！\n'
write_txt('/jd/scripts/log/'+str(time.strftime('%m-%d', time.localtime()))+'.txt',logging)
for i in range(startshop,startshop+run_times):
    try:
        runtimes+=1
        tellme=tellme+1
        url = "https://shop.m.jd.com/?shopId=" + str(i)
        # 模仿浏览器的headers
        headers = {
            "user-agent": "Mozilla/5.0 (Linux; Android 9; STF-AL10; HMSCore 6.4.0.311; GMSCore 19.6.29) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 HuaweiBrowser/11.0.4.371 Mobile Safari/537.36",
            "Host" : "shop.m.jd.com",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            'upgrade-insecure-requests': '1',
            'accept-encoding': 'gzip, deflate',
            'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
            'sec-fetch-site': 'none',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-dest': 'document',
            'cookie':'visitkey='+visitkey,
        }
        # 调用get方法，传入参数，返回结果集
        resp = requests.get(url,headers=headers,timeout=(30.05, 60.06))
        jsonD = json.dumps(resp.text)
        vid=r"\d+"
        a=re.search(vid,jsonD, re.M|re.I)
        if jsonD.find('Authorization Required')!=-1:
            print('黑了，偿试visitkey='+str(int(visitkey)+1)+'看看')
            visitkey=str(int(visitkey)+1)
            blacktimes +=1
            if blacktimes==30 : break
            continue
        if tellme % 20==0 :         #20次休息2-5秒
            #print('休息几秒',jsonD[:40])
            time.sleep(int(5*random.random()+5))
        if a==None:
            continue
        url2='https://wq.jd.com/shopbranch/GetUrlSignDraw?channel=1&venderId='+str(a.group(0))
        header2={
            "user-agent": "Mozilla/5.0 (Linux; Android 9; STF-AL10; HMSCore 6.4.0.311; GMSCore 19.6.29) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 HuaweiBrowser/11.0.4.371 Mobile Safari/537.36",
            "Host" : "wq.jd.com",
            "Accept": "*/*",
            "Accept-Language": "zh-cn",
            "Accept-encoding":"gzip, deflate",
            'accept-language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
            'referer': 'https://shop.m.jd.com/?shopId='+str(i),
            'cookie':'visitkey='+visitkey,
        }

        resp2 = requests.get(url2,headers=header2,timeout=(30.05, 60.06))
        json2 = json.dumps(resp2.text)
        tokenid=re.search("token\%3D[a-zA-Z0-9]+",json2, re.M|re.I)
        if json2.find('Authorization Required')!=-1:
            print('又黑了，偿试visitkey='+str(int(visitkey)+1)+'看看')
            visitkey=str(int(visitkey)+1)
            blacktimes +=1
            if blacktimes==30 : break
            continue
        if tokenid==None:
            if tellme % 1000==0 :
                logging=str(time.strftime('%H:%M:%S', time.localtime()))+':第'+str(i)+'个店铺,还是没有\n'
                write_txt('/jd/scripts/log/'+str(time.strftime('%m-%d', time.localtime()))+'.txt',logging)
            continue
        tokenid=tokenid.group(0)[-32:]
        print('发现有，存入'+str(i))
        shopaddr='https://shop.m.jd.com/?shopid='+str(i)

        url3='https://api.m.jd.com/api?appid=interCenter_shopSign&t=1648774694000&loginType=2&functionId=interact_center_shopSign_getActivityInfo&body={%22token%22:%22'+tokenid+'%22,%22venderId%22:%22%22}&jsonp=jsonp1000'
        header3={
            'accept':'*/*',
            'accept-encoding': 'gzip, deflate, br',
            'user-agent': 'jdapp;iPhone;10.4.4;;;M/5.0;appBuild/167991;jdSupportDarkMode/0;ef/1;ep/%7B%22ciphertype%22%3A5%2C%22cipher%22%3A%7B%22ud%22%3A%22CzOzCzc0YwVtDwHuDJuyY2GyEJC0EJG5ZNCyYtK3CJU0YzU3YtCyYm%3D%3D%22%2C%22sv%22%3A%22CJCkCy4n%22%2C%22iad%22%3A%22HJC0Dtq0C0OjENKzGy00HJHNBJrPGJCjHtPLDzLPDUSmDJHL%22%7D%2C%22ts%22%3A1648774693%2C%22hdid%22%3A%22JM9F1ywUPwflvMIpYPok0tt5k9kW4ArJEU3lfLhxBqw%3D%22%2C%22version%22%3A%221.0.3%22%2C%22appname%22%3A%22com.360buy.jdmobile%22%2C%22ridx%22%3A-1%7D;Mozilla/5.0 (iPhone; CPU iPhone OS 13_3_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148;supportJDSHWK/1;',
            'accept-language': 'zh-cn',
            'referer': 'https://h5.m.jd.com/babelDiy/Zeus/2PAAf74aG3D61qvfKUM5dxUssJQ9/index.html?token=DFBDE9D4A0EF038284396FDA8535D20C&sid=4c8be78390c78d263a9486e98dcacafw&un_area=19_1666_1669_52081',
            'cookie':'visitkey='+visitkey,
        }

        resp3 = requests.get(url3,headers=header3,timeout=(30.05, 60.06))
        json3 = json.dumps(resp3.text)
        res='(?<=\().*(?=\))'
        jsonE=re.search(res,json3,re.M|re.I).group(0).replace('\\','')
        infor=''
        rulelist=[1,1]
        rulelist=json.loads(jsonE).get('data').get('continuePrizeRuleList') #奖励规则 是列表对象
        listnum=len(rulelist)
        for ax in range(listnum):
            days=rulelist[ax].get('level')     #每份奖励的天数
            prizelist=rulelist[ax].get('prizeList')   #每份奖励的详细规则列表
            for k in range(len(prizelist)):
                if (prizelist[k].get('type')==4):   #type是4是豆子 14是红包
                    infor=infor+f'每{days}天奖京豆:{prizelist[k].get("discount")}个,'
                if (prizelist[k].get('type')==14):
                    infor=infor+f'每{days}天奖红包:{prizelist[k].get("discount")}分。。。,'
                #print (infor)
        #运行完没有期望奖励，则不保存，继续下个店铺
        if(infor==''):
            print('有活动，但是不是豆子红包，跳过继续下一个店铺。')
            continue

        ns[j]=[shopaddr,tokenid,time.strftime('%H:%M:%S', time.localtime())]
        successTimes+=1
        j=j+1

        """ data = openpyxl.load_workbook('data.xlsm',keep_vba=True)
        table = data.worksheets[0]
        table = data.active
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得列数
        for r in range(len(ns)):
            for c in range(len(ns[0])):
                table.cell(r + nrows+1, c + 1).value = ns[r][c]
        data.save('data.xlsm') """
        temp=str(time.strftime('%m-%d-%H:%M:%S', time.localtime()))+';'+ns[0][0] +';'+str(tokenid)+';'+infor+'\n'
        write_txt('/jd/scripts/data.txt',temp)
        j=0
        temp=''
        ns = [[0 for x in range(3)] for y in range(1)]
    except BaseException as e:
        print('停止，正在保存文件',e)

        logging=str(time.strftime('%H:%M:%S', time.localtime()))+f':出现错误停止\n'+f'{e}'
        write_txt('/jd/scripts/log/'+str(time.strftime('%m-%d', time.localtime()))+'.txt',logging)
        break
#最后把上次执行到的店铺号写入文件中
alter("/jd/scripts/main.py", "startshop="+str(startshop), "startshop="+str(i))
alter("/jd/scripts/main.py", "visitkey='"+modikey+"'", "visitkey='"+visitkey+"'")
###############################
#记录运行时间
end_time = time.time()
run_time = round(end_time-startTime)
hour = run_time//3600
minute = (run_time-3600*hour)//60
second = run_time-3600*hour-60*minute
logging=str(time.strftime('%H:%M:%S', time.localtime()))+f':全部运行完毕，耗时：{hour}小时{minute}分钟{second}秒,共找了{runtimes}个店铺，找到目标{successTimes}个,更换身份{blacktimes}次\n'
write_txt('/jd/scripts/log/'+str(time.strftime('%m-%d', time.localtime()))+'.txt',logging)


